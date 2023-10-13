import os
import xlrd
from pypdf import PdfReader
from zipfile import ZipFile
from openpyxl import load_workbook


def test_is_files_in_archive():
    # ARRANGE
    root_folder = os.getcwd()
    files_folder = os.path.join(root_folder, "resources")
    zip_file = os.path.join(root_folder, "tmp", "files.zip")
    files = os.listdir(files_folder)

    # ACT
    with ZipFile(zip_file) as zf:
        # ASSERT
        assert files == zf.namelist()


def test_is_same_txt_file():
    # ARRANGE
    root_folder = os.getcwd()
    files_folder = os.path.join(root_folder, "resources")
    zip_file = os.path.join(root_folder, "tmp", "files.zip")
    test_file_name = "test_file1.txt"
    searched_text = test_file_name
    test_file_size = os.path.getsize(os.path.join(files_folder, test_file_name))

    # ACT
    with ZipFile(zip_file) as zf:
        text_file = zf.read(test_file_name).decode(encoding='utf-8')
        # ASSERT
        assert test_file_size == zf.getinfo(test_file_name).file_size
        assert searched_text in text_file


def test_is_same_pdf_file():
    # ARRANGE
    root_folder = os.getcwd()
    files_folder = os.path.join(root_folder, "resources")
    zip_file = os.path.join(root_folder, "tmp", "files.zip")
    test_file_name = "test_file2.pdf"
    searched_text = test_file_name
    test_file_size = os.path.getsize(os.path.join(files_folder, test_file_name))

    # ACT
    with ZipFile(zip_file) as zf:
        pdf_file = PdfReader(zf.open(test_file_name))
        # ASSERT
        assert test_file_size == zf.getinfo(test_file_name).file_size
        assert searched_text in pdf_file.pages[0].extract_text()


def test_is_same_xls_file():
    # ARRANGE
    root_folder = os.getcwd()
    files_folder = os.path.join(root_folder, "resources")
    zip_file = os.path.join(root_folder, "tmp", "files.zip")
    test_file_name = "test_file3.xls"
    searched_text = test_file_name
    test_file_size = os.path.getsize(os.path.join(files_folder, test_file_name))

    # ACT
    with ZipFile(zip_file) as zf:
        xls_file = xlrd.open_workbook(file_contents=zf.read(test_file_name))
        # ASSERT
        assert test_file_size == zf.getinfo(test_file_name).file_size
        assert searched_text == xls_file.sheet_by_index(0).cell_value(0, 0)


def test_is_same_xlsx_file():
    # ARRANGE
    root_folder = os.getcwd()
    files_folder = os.path.join(root_folder, "resources")
    zip_file = os.path.join(root_folder, "tmp", "files.zip")
    test_file_name = "test_file4.xlsx"
    searched_text = test_file_name
    test_file_size = os.path.getsize(os.path.join(files_folder, test_file_name))

    # ACT
    with ZipFile(zip_file) as zf:
        xlsx_file = load_workbook(zf.open(test_file_name, 'r'))
        # ASSERT
        assert test_file_size == zf.getinfo(test_file_name).file_size
        assert searched_text == xlsx_file.active.cell(3, 2).value
